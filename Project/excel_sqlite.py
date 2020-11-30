import openpyxl as op
from openpyxl.styles import Font, Fill
import sqlite3

ConStr="EmpExcelData.db"
print('OPENING WORKBOOK Employee_Data......')

wb = op.load_workbook('Employee_Data.xlsx')  # Open workbook Employee Data
sheet = wb.get_sheet_by_name('Employee data')  # Gets Sheet by sheetname which is Employee data
sheet.title='emp_data'                 # Sets title to emp_data
mysheet=wb.active
print('Active sheet {}'.format(mysheet))
print('There are {} rows and {} columns in Employee Data'.format(sheet.max_row,sheet.max_column))   # Highest row and column of the sheet
sheet['F1']='Job Hours' # Column heading changes from Job_Time to Job Hours

def create_table():
    conn = sqlite3.connect(ConStr)
    cr=conn.cursor()
    cr.execute('''CREATE TABLE Employee
                 (Emp_id INTEGER PRIMARY KEY, Emp_name TEXT, DOB TEXT,
         Job_Cat INTEGER, Salary INTEGER, Job_Hour TEXT,Project INTEGER,Prev_Exp INTEGER,Gender text)''') # Create table Employee

    conn.commit()
    conn.close()


def insert_into_table(sheet):
    conn = sqlite3.connect(ConStr)
    cr=conn.cursor()
    for i in range(2, sheet.max_row):
        for j in range(1,sheet.max_column+1):
            Emp_id,Emp_name,DOB,Job_Cat,Salary=sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=4).value,sheet.cell(row=i,column=5).value
            Job_Time,Project,Prev_Exp,Gender=sheet.cell(row=i,column=6).value,sheet.cell(row=i,column=7).value,sheet.cell(row=i,column=8).value,sheet.cell(row=i,column=9).value
            cr.execute('''INSERT OR IGNORE INTO Employee
             (Emp_id,Emp_name,DOB,Job_Cat,Salary,Job_Hour,Project,Prev_Exp,Gender)
            VALUES ( ?, ?, ?, ?, ?, ? ,?,?,?)''',
            ( Emp_id,Emp_name,DOB,Job_Cat,int(Salary),Job_Time,Project,Prev_Exp,Gender)) # Inserts all Excel data into Database ExcelData

    conn.commit()
    conn.close()

def delete_rows():
    conn = sqlite3.connect(ConStr)
    cr=conn.cursor()
    cr.execute('''DELETE FROM Employee WHERE Emp_id>480''')      #As there are 490 rows in table, all rows which are greater then 480 are deleted
    conn.commit()
    conn.close()


def update_table():
    conn = sqlite3.connect(ConStr)
    cr=conn.cursor()
    cr.execute('''UPDATE Employee
                SET Salary = 80000, Prev_Exp= Prev_Exp +1
                WHERE Emp_name LIKE 's%' and Prev_Exp>20;''')    #Updates the table by setting salary as 80000, and incrementing there Previous Experience
                                                          #for employee with name that starts and ends with s and a respectively and Experience>20
    conn.commit()
    conn.close()

def select_cols_based_on_project():
    conn = sqlite3.connect(ConStr)
    cr=conn.cursor()
    REC=cr.execute('''SELECT Emp_id,Emp_name,Salary,
    case
        when date(dob, '+' ||
            strftime('%Y', 'now') - strftime('%Y', dob) ||
            ' years') >= date('now')
        then strftime('%Y', 'now') - strftime('%Y', dob)
        else strftime('%Y', 'now') - strftime('%Y', dob) - 1
    end
    as Age,
	project,Prev_Exp
    FROM Employee WHERE PROJECT >150 and Prev_Exp >=15 and Salary>=80000  ORDER BY Project Desc;''') #Selects Emp_id,name,salary,age,project,Prev_Exp from Database

    display_records(REC)
    conn.commit()
    conn.close()

def display_records(rec):
    print('Employee Id, '+'Employee Name, '+'Salary, '+'Age, '+'Project, '+'Prev Experience')  #Display Records stored in  database
    records=rec.fetchall()
    for r in range(len(records)):
        print(str(records[r][0])+' ,       '+str(records[r][1])+'  ,      '+str(records[r][2])+'  , '+str(records[r][3])+',   '+str(records[r][4])+'  ,    '+str(records[r][5]))
    storing_newdata_excel(records)



def storing_newdata_excel(records):
    wb.create_sheet(title='Updated Employee Data')
    sheet=wb.get_sheet_by_name('Updated Employee Data')

    sheet['A1'].font= Font(size=14, italic=True,bold=True)      #Styling Excel Sheets
    sheet['A1'] = 'EMPLOYEE\'S DATA FROM DATABASE'

    sheet['A2'],sheet['B2'],sheet['C2'],sheet['D2'],sheet['E2'],sheet['F2']='Emp Id ','Emp Name ','Salary ','Age ','Project ','Prev Experience'

    i=0
    for rowNum in range(3,len(records)+3):
        sheet.cell(row=rowNum,column=1).value,sheet.cell(row=rowNum,column=2).value,sheet.cell(row=rowNum,column=3).value=records[i][0],records[i][1],records[i][2]
        sheet.cell(row=rowNum,column=4).value,sheet.cell(row=rowNum,column=5).value,sheet.cell(row=rowNum,column=6).value=records[i][3],records[i][4],records[i][5]
        i+=1

    wb.save('NewEmp_Data.xlsx')                          #Save WORKBOOK


# Function Calls
create_table()
insert_into_table(sheet)
delete_rows()
update_table()
select_cols_based_on_project()
